import json
import os
from openpyxl import Workbook
import pandas as pd
from datetime import datetime
import openpyxl
from pathlib import Path
import logging
import re


class Excel_parser:
    def __init__(self):
        self.symbol_path = ""
        self.value_path = ""
        self.root_dir = Path(__file__).parent.parent

        # 設定 logging
        logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

    def parse_filename(self, filename):
        """解析檔名，提取遊戲名稱與回合數"""
        match = re.search(r"(.+)_round_(\d+)-", filename)
        if match:
            return match.group(1), int(match.group(2))
        return None, None

    def extract_value(self, value_str):
        """提取數值，支持浮點數，並去除千分位逗號"""
        try:
            return float(value_str.replace(',', '')) if value_str else 0
        except ValueError:
            logging.warning(f"數值轉換失敗：{value_str}")
            return 0

    def compute_rtp(self, data_folder, game_type, output_excel="rtp_results.xlsx"):

        Path(os.path.join(Path(__file__).parent.parent, "excel")).mkdir(exist_ok=True)
        output_excel = os.path.join(Path(__file__).parent.parent, 'excel', f'{game_type}_rtp_results.xlsx')


        """計算 RTP 並輸出 Excel 檔案"""
        rounds_data = {}

        if not os.path.exists(data_folder):
            logging.error(f"資料夾不存在：{data_folder}")
            return None
        
        # 遍歷 JSON 檔案，提取每個回合的數據
        for file in sorted(os.listdir(data_folder)):
            if not file.endswith(".json"):
                continue

            game_name, round_number = self.parse_filename(file)
            if game_name is None:
                logging.warning(f"無法解析檔名：{file}")
                continue

            file_path = os.path.join(data_folder, file)
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
            except (json.JSONDecodeError, FileNotFoundError) as e:
                logging.error(f"讀取失敗 {file}: {e}")
                continue

            balance = self.extract_value(data.get("玩家剩餘金額", {}).get("value", "0"))
            bet = self.extract_value(data.get("押注金額", {}).get("value", "0"))
            win = self.extract_value(data.get("玩家贏分", {}).get("value", "0"))  # 優先使用 JSON 提供的贏分

            if round_number not in rounds_data:
                rounds_data[round_number] = {
                    "end_balance": balance, 
                    "bet": bet,
                    "max_win": win  # 記錄該回合的最大贏分
                }
            else:
                rounds_data[round_number]["end_balance"] = balance
                rounds_data[round_number]["max_win"] = max(rounds_data[round_number]["max_win"], win)  # 取最大值

        # 計算 RTP
        results = []
        total_bet, total_win = 0, 0
        previous_balance = None

        for rnd in sorted(rounds_data.keys()):
            data = rounds_data[rnd]

            # 計算贏分，優先使用 JSON 記錄的最大贏分
            if data["max_win"] > 0:
                win = data["max_win"]
            else:
                # 若 JSON 無提供贏分，則使用餘額變化計算
                win = max(0, data["end_balance"] - previous_balance) if previous_balance is not None else 0

            total_bet += data["bet"]
            total_win += win

            try:
                rtp = win / data["bet"] if data["bet"] > 0 else 0
            except ZeroDivisionError:
                rtp = 0

            results.append([rnd, data["bet"], win, f"{rtp:.2%}"])
            logging.info(f"Round {rnd}: Bet = {data['bet']}, Win = {win}, RTP = {rtp:.2%}")

            # 更新上一回合的餘額
            previous_balance = data["end_balance"]

        overall_rtp = total_win / total_bet if total_bet > 0 else 0
        logging.info(f"Overall RTP: {overall_rtp:.2%}")

        # 存入 Excel
        df = pd.DataFrame(results, columns=["Round", "Bet", "Win", "RTP"])
        df.to_excel(output_excel, index=False)
        logging.info(f"RTP 結果已儲存至 {output_excel}")

        return overall_rtp


    def json_to_excel(self, game_type, game_state):

        Path(os.path.join(self.root_dir, "excel")).mkdir(exist_ok=True)
        save_path = os.path.join(self.root_dir, 'excel', f'{game_type}_{game_state}.xlsx')
        self.symbol_path = os.path.join(self.root_dir, 'output', game_type,'symbols')
        self.value_path = os.path.join(self.root_dir, 'output', game_type, 'numerical')
        # 建立資料
        excel = {
            "遊戲名稱": [],
            "測試時間": [],
            "遊戲狀態": [],
        }

        excel_value = {}
        excel_symbol = {}

        #讀取數值檔案
        #建立鍵值
        value_file_list = os.listdir(self.value_path)
        #value_file_list.sort(key=lambda x: int(x.split('_')[-1].split('.')[0]))
        
        # 排序 value_file_list
        value_file_list.sort(key=lambda x: tuple(map(int, x.split('_')[-1].split('.')[0].split('-'))))

        for file_name in value_file_list:
            excel['遊戲名稱'].append(game_type)
            excel['測試時間'].append("")
            excel['遊戲狀態'].append(game_state)

            # 檢查是否為 JSON 檔案
            if file_name.endswith(".json"):
                file_path = os.path.join(self.value_path, file_name)
                with open(file_path, "r", encoding="utf-8") as file:
                    # 解析 JSON 檔案
                    json_data = json.load(file)
                    # 建立鍵值
                    for key in json_data:
                        if not key in excel_value:
                            excel_value[key] = []

        #輸入數值
        for file_name in value_file_list:
            # 檢查是否為 JSON 檔案
            if file_name.endswith(".json"):
                file_path = os.path.join(self.value_path, file_name)
                with open(file_path, "r", encoding="utf-8") as file:
                    # 解析 JSON 檔案
                    json_data = json.load(file)
                    for key in excel_value:
                        if not key in json_data:
                            excel_value[key].append("")
                        else:
                            excel_value[key].append(json_data[key]['value'])


        # 讀取盤面檔案
        # 建立鍵值
        symbol_file_list = os.listdir(self.symbol_path)
        # 修改排序逻辑，处理文件名中包含非数字字符的情况
        symbol_file_list.sort(key=lambda x: tuple(map(int, x.split('_')[-1].split('.')[0].split('-'))))
        for file_name in symbol_file_list:
            # 檢查是否為 JSON 檔案
            if file_name.endswith(".json"):
                file_path = os.path.join(self.symbol_path, file_name)
                with open(file_path, "r", encoding="utf-8") as file:
                    # 解析 JSON 檔案
                    json_data = json.load(file)
                    # 建立鍵值
                    for grid in json_data:
                        symbol_pos = f"C{grid['value'][1]}R{grid['value'][0]}"
                        # 建立鍵值
                        if not symbol_pos in excel_symbol:
                            excel_symbol[symbol_pos] = []

        # 輸入數值
        for file_name in symbol_file_list:
            # 檢查是否為 JSON 檔案
            if file_name.endswith(".json"):
                file_path = os.path.join(self.symbol_path, file_name)
                with open(file_path, "r", encoding="utf-8") as file:
                    # 解析 JSON 檔案
                    json_data = json.load(file)

                    #problem
                    for key in excel_symbol:
                        found = False
                        for grid in json_data:
                            symbol_pos = f"C{grid['value'][1]}R{grid['value'][0]}"
                            if symbol_pos == key:
                                found = True
                                excel_symbol[key].append(grid['key'])
                                break
                        if not found:
                            excel_symbol[key].append("")


        excel.update(excel_symbol)
        excel.update(excel_value)

        # 建立 DataFrame
        df = pd.DataFrame(excel)

        # 儲存檔案
        df.to_excel(save_path)
        print("檔案已成功儲存！")

    def get_file_creation_time(self, file_path):
        try:
            creation_timestamp = os.path.getctime(file_path)
            creation_time = datetime.fromtimestamp(creation_timestamp).strftime("%Y-%m-%d %H:%M:%S")
            return creation_time
        except Exception as e:
            print(f"Error getting creation time for {file_path}: {e}")
            return None

    def fill_creation_times_by_index(self, folder_path, excel_path, output_excel):
        if not os.path.exists(folder_path):
            print(f"Folder not found: {folder_path}")
            return

        if not os.path.exists(excel_path):
            print(f"Excel file not found: {excel_path}")
            return

        # Get all JSON files in the folder
        json_files = sorted([f for f in os.listdir(folder_path) if f.endswith(".json")])
        if not json_files:
            print("No JSON files found in the specified folder.")
            return

        # Load the Excel workbook
        workbook = openpyxl.load_workbook(excel_path)
        sheet = workbook.active

        # Find the column for "測試時間"
        headers = [cell.value for cell in sheet[1]]
        if "測試時間" not in headers:
            print("Column '測試時間' not found in Excel file.")
            return

        time_column_index = headers.index("測試時間") + 1

        # Fill creation times based on the index
        for row in range(2, sheet.max_row + 1):
            index = row - 2  # Index starts from 0 for JSON files
            if index < len(json_files):
                json_file = json_files[index]
                json_path = os.path.join(folder_path, json_file)
                creation_time = self.get_file_creation_time(json_path)
                if creation_time:
                    sheet.cell(row=row, column=time_column_index, value=creation_time)

        # Save the updated Excel file
        workbook.save(output_excel)
        print(f"Updated Excel file saved to {output_excel}")


'''
if __name__ == "__main__":
    ex = Excel_parser()
    ex.json_to_excel("golden","base")
'''